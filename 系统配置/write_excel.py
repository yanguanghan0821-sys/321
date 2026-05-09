#!/usr/bin/env python3
"""
桌面竞彩数据库脚本 — 自动写入/更新桌面Excel
用法：
  python3 write_excel.py add_matches <json_data>       # 添加比赛数据到历史数据库
  python3 write_excel.py add_review <json_data>         # 添加复盘记录
  python3 write_excel.py read_today                     # 读取今日推荐（供分析用）
  python3 write_excel.py stats                          # 输出统计数据
"""
import sys, json, os, openpyxl
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB = "/mnt/c/Users/Administrator/Desktop/竞彩数据系统_v3.xlsx"
FALLBACK_DB = "/tmp/竞彩数据系统_v3.xlsx"

def get_db():
    """获取可写的数据库路径"""
    if os.access(DB, os.W_OK) or (os.path.exists(DB) and os.access(os.path.dirname(DB), os.W_OK)):
        return DB
    return FALLBACK_DB

def ensure_sheet(wb, name):
    """确保工作表存在"""
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        return ws
    return wb[name]

### 样式定义 ###
F = Font(name='Arial', color='E8EEF8', size=10)
FH = Font(name='Arial', bold=True, color='FFCB45', size=11)
FT = Font(name='Arial', bold=True, color='FFCB45', size=14)
R1 = PatternFill(start_color='0A0F18', end_color='0A0F18', fill_type='solid')
R2 = PatternFill(start_color='0D1322', end_color='0D1322', fill_type='solid')
HD = PatternFill(start_color='0F1622', end_color='0F1622', fill_type='solid')
B = Border(left=Side(style='thin',color='1C2942'),right=Side(style='thin',color='1C2942'),
           top=Side(style='thin',color='1C2942'),bottom=Side(style='thin',color='1C2942'))
C = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_sheet(ws, headers, widths):
    """应用样式到工作表"""
    cols = len(headers)
    # 先取消所有合并单元格
    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = FH; c.fill = HD; c.alignment = C; c.border = B
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 已有数据行应用样式
    for ri in range(2, ws.max_row + 1):
        for ci in range(1, cols + 1):
            try:
                c = ws.cell(row=ri, column=ci)
                if not c.font or c.font == Font():
                    c.font = F; c.fill = R1 if ri%2==0 else R2; c.alignment = C; c.border = B
            except AttributeError:
                pass


### 添加比赛到历史数据库 ###
def add_matches(matches_data):
    """批量添加比赛数据到历史数据库工作表"""
    db = get_db()
    try:
        wb = openpyxl.load_workbook(db)
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "历史数据库"
        headers = ["日期","联赛","轮次","主队","客队","全场比分","半场比分",
                    "B365主胜","B365平","B365客胜","B365盘口","B365主水","B365客水","B365返还率",
                    "平博主胜","平博平","平博客胜","平博盘口","平博主水","平博客水",
                    "联赛级别","冷门标记","结果分析"]
        widths = [16,10,8,16,16,10,8,10,10,10,12,8,8,10,10,10,10,12,8,8,10,12,20]
        style_sheet(ws, headers, widths)
        wb.save(db)

    ws = ensure_sheet(wb, "历史数据库")
    # 确保表头存在 - 先处理可能的合并单元格
    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))
    try:
        header_exists = ws.max_row >= 1 and ws.cell(row=1, column=1).value == "日期"
    except AttributeError:
        header_exists = False
    if not header_exists:
        headers = ["日期","联赛","轮次","主队","客队","全场比分","半场比分",
                    "B365主胜","B365平","B365客胜","B365盘口","B365主水","B365客水","B365返还率",
                    "平博主胜","平博平","平博客胜","平博盘口","平博主水","平博客水",
                    "联赛级别","冷门标记","结果分析"]
        widths = [16,10,8,16,16,10,8,10,10,10,12,8,8,10,10,10,10,12,8,8,10,12,20]
        style_sheet(ws, headers, widths)

    added = 0
    for m in matches_data:
        row = ws.max_row + 1
        data = [
            m.get("date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            m.get("league", ""),
            m.get("round", ""),
            m.get("home", ""),
            m.get("away", ""),
            m.get("score", ""),
            m.get("ht_score", ""),
            m.get("b365_win", ""), m.get("b365_draw", ""), m.get("b365_lose", ""),
            m.get("b365_hdcp", ""), m.get("b365_h_water", ""), m.get("b365_a_water", ""),
            m.get("b365_return", ""),
            m.get("pb_win", ""), m.get("pb_draw", ""), m.get("pb_lose", ""),
            m.get("pb_hdcp", ""), m.get("pb_h_water", ""), m.get("pb_a_water", ""),
            m.get("level", ""), m.get("upset", ""), m.get("analysis", "")
        ]
        for ci, v in enumerate(data, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = F; c.fill = R1 if row%2==0 else R2; c.alignment = C; c.border = B
        added += 1

    wb.save(db)
    print(f"✅ 已添加 {added} 场比赛到历史数据库 [{db}]")


### 添加复盘 ###
def add_review(review_data):
    """添加复盘记录"""
    db = get_db()
    try:
        wb = openpyxl.load_workbook(db)
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "每日复盘"
        headers = ["日期","场次","联赛","主队","客队","推荐玩法","赔率","比分","结果","命中","分析"]
        widths = [12,10,10,14,14,12,10,10,8,8,24]
        style_sheet(ws, headers, widths)
        wb.save(db)

    ws = ensure_sheet(wb, "每日复盘")
    if ws.max_row < 1 or ws.cell(row=1, column=1).value != "日期":
        headers = ["日期","场次","联赛","主队","客队","推荐玩法","赔率","比分","结果","命中","分析"]
        widths = [12,10,10,14,14,12,10,10,8,8,24]
        style_sheet(ws, headers, widths)

    r = review_data
    row = ws.max_row + 1
    data = [
        r.get("date", date.today().strftime("%Y-%m-%d")),
        r.get("match_no", ""), r.get("league", ""),
        r.get("home", ""), r.get("away", ""),
        r.get("pick", ""), r.get("odds", ""),
        r.get("score", ""), r.get("result", ""),
        r.get("hit", ""), r.get("analysis", "")
    ]
    for ci, v in enumerate(data, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = F; c.fill = R1 if row%2==0 else R2; c.alignment = C; c.border = B

    wb.save(db)
    print(f"✅ 复盘已添加: {r.get('home','')} {r.get('score','')} {r.get('away','')}")


### 读取今日推荐 ###
def read_today():
    """读取今日推荐数据（JSON输出供调用）"""
    db = get_db()
    try:
        wb = openpyxl.load_workbook(db)
    except:
        print(json.dumps({"error": "无法打开数据库"}))
        return

    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append([str(v) if v is not None else "" for v in row])
        result[sheet_name] = rows

    print(json.dumps(result, ensure_ascii=False))


### 统计信息 ###
def stats():
    """输出数据库统计"""
    db = get_db()
    try:
        wb = openpyxl.load_workbook(db)
    except:
        print(json.dumps({"error": "无法打开数据库"}))
        return

    stats_data = {}
    for name in wb.sheetnames:
        ws = wb[name]
        stats_data[name] = {
            "行数": ws.max_row - 1 if ws.max_row > 1 else 0,
            "列数": ws.max_column
        }

    print(json.dumps(stats_data, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法:")
        print("  python3 write_excel.py add_matches '<json>'")
        print("  python3 write_excel.py add_review '<json>'")
        print("  python3 write_excel.py read_today")
        print("  python3 write_excel.py stats")
        sys.exit(1)

    action = sys.argv[1]

    if action == "add_matches":
        data = json.loads(sys.argv[2]) if len(sys.argv) > 2 else []
        # 支持单条或批量
        if isinstance(data, dict):
            data = [data]
        add_matches(data)

    elif action == "add_review":
        data = json.loads(sys.argv[2]) if len(sys.argv) > 2 else {}
        add_review(data)

    elif action == "read_today":
        read_today()

    elif action == "stats":
        stats()

    else:
        print(f"未知操作: {action}")
        sys.exit(1)
