#!/usr/bin/env python3
"""
每日竞彩同步脚本 — 一站式写入总表Excel + 竞彩分析库
用法：
  # 写入当日数据（给agent调用）
  python3 daily_sync.py all <日期> <比赛数据JSON> <分析数据JSON>

  # 单独写入某部分
  python3 daily_sync.py matches <日期> '<比赛数据JSON>'
  python3 daily_sync.py analysis <日期> '<分析结果JSON>'
  python3 daily_sync.py review <日期> '<复盘数据JSON>'

比赛数据JSON格式：
  [{"league":"英超","home":"利物浦","away":"阿森纳","win_odds":"1.72","draw_odds":"3.50","lose_odds":"4.00",
    "hdcp":"半球","h_water":"0.95","a_water":"0.90","time":"18:00","match_no":"周一001"}]

分析结果JSON格式：
  {"main_picks":[{"priority":"⭐⭐","match":"利物浦 vs 阿森纳","pick":"主胜","odds":"1.72","confidence":"7/10","reason":"主场强势"}],
   "high_odds":[...], "parlays":[...],
   "lotto14":[{"no":"1","league":"欧冠","home":"阿森纳","away":"马竞","win":"1.64","draw":"4.03","lose":"5.60","pick":"3","dan":"是"}],
   "ren9":{"desc":"5胆4双=16注32元","picks":[{"type":"⭐01","match":"阿森纳vs马竞","pick":"3"}]},
   "model_update":["意甲平率+2%"]}
"""
import json, sys, os, subprocess
from datetime import date, datetime

DESKTOP = "/mnt/c/Users/Administrator/Desktop"
MASTER_DB = os.path.expanduser("~/.hermes/data/Aii竞彩数据总表.xlsx")
DESKTOP_DB = f"{DESKTOP}/Aii竞彩数据总表.xlsx"
OBSIDIAN_VAULT = f"{DESKTOP}/足球数据分析库"  # 2026-05-06 从'竞彩分析库'改为'足球数据分析库'以匹配skill最新指向
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))

def today():
    return date.today().strftime("%Y-%m-%d")

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def run_script(script_name, *args):
    """运行指定脚本"""
    script_path = f"{SCRIPTS_DIR}/{script_name}"
    if not os.path.exists(script_path):
        script_path = f"{os.path.expanduser('~')}/.hermes/scripts/{script_name}"
    cmd = ["/usr/bin/python3", script_path] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"⚠️ {script_name} 错误: {result.stderr[:200]}")
        return False
    print(result.stdout.strip())
    return True


def write_to_master_excel(pred_date, matches_data, analysis_data):
    """写入总表Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    F = Font(name='Arial', color='E8EEF8', size=10)
    FH = Font(name='Arial', bold=True, color='FFCB45', size=11)
    R1 = PatternFill(start_color='0A0F18', end_color='0A0F18', fill_type='solid')
    R2 = PatternFill(start_color='0D1322', end_color='0D1322', fill_type='solid')
    B = Border(left=Side(style='thin',color='1C2942'),right=Side(style='thin',color='1C2942'),
               top=Side(style='thin',color='1C2942'),bottom=Side(style='thin',color='1C2942'))
    C = Alignment(horizontal='center', vertical='center', wrap_text=True)

    try:
        wb = openpyxl.load_workbook(MASTER_DB)
    except:
        print(f"⚠️ 无法打开总表 {MASTER_DB}")
        return False

    ws = wb["竞彩数据"]
    changed = False

    # 建立去重键集合
    seen = set()
    for r in range(2, ws.max_row + 1):
        d = safe_str(ws.cell(r, 1).value)
        h = safe_str(ws.cell(r, 6).value)
        a = safe_str(ws.cell(r, 7).value)
        t = safe_str(ws.cell(r, 2).value)
        seen.add(f"{d}|{h}|{a}|{t}")

    # 写入比赛数据
    if matches_data:
        added = 0
        for m in matches_data:
            key = f"{pred_date}|{m.get('home','')}|{m.get('away','')}|竞彩"
            if key in seen: continue
            seen.add(key)
            row = ws.max_row + 1
            data = [
                pred_date, "竞彩", m.get("league",""), "", m.get("match_no",""),
                m.get("home",""), m.get("away",""), "", "",
                m.get("win_odds",""), m.get("draw_odds",""), m.get("lose_odds",""),
                m.get("hdcp",""), m.get("h_water",""), m.get("a_water",""), "",
                "", "", "", "", "", "", "", "", "",
                "", "", "", "", "自抓取"
            ]
            for ci, v in enumerate(data, 1):
                c = ws.cell(row=row, column=ci, value=v)
                try: c.font = F; c.fill = R1 if row%2==0 else R2; c.alignment = C; c.border = B
                except: pass
            added += 1
            changed = True
        print(f"  📊 比赛数据 +{added} 行")

    # 写入14场数据
    lotto14 = analysis_data.get("lotto14", []) if analysis_data else []
    if lotto14:
        added = 0
        for m in lotto14:
            key = f"{pred_date}|{m.get('home','')}|{m.get('away','')}|14场"
            if key in seen: continue
            seen.add(key)
            row = ws.max_row + 1
            ws.cell(row, 1, value=pred_date)
            ws.cell(row, 2, value="14场")
            ws.cell(row, 3, value=m.get("league",""))
            ws.cell(row, 4, value=m.get("no",""))
            ws.cell(row, 5, value=m.get("home",""))
            ws.cell(row, 6, value=m.get("away",""))
            ws.cell(row, 9, value=m.get("pick",""))
            ws.cell(row, 10, value=m.get("win",""))
            ws.cell(row, 11, value=m.get("draw",""))
            ws.cell(row, 12, value=m.get("lose",""))
            ws.cell(row, 36, value=m.get("pick",""))
            ws.cell(row, 37, value="胆" if m.get("dan","")=="是" else "")
            ws.cell(row, 44, value="自抓取")
            for c in range(1, 45):
                cell = ws.cell(row, c)
                try: cell.font = F; cell.fill = R1 if row%2==0 else R2; cell.alignment = C; cell.border = B
                except: pass
            added += 1
            changed = True
        print(f"  📊 14场数据 +{added} 行")

    # 写入推荐分析数据
    picks = analysis_data.get("main_picks", []) if analysis_data else []
    if picks:
        added = 0
        for p in picks:
            key = f"{pred_date}|{p.get('match','')}|推荐"
            if key in seen: continue
            seen.add(key)
            row = ws.max_row + 1
            ws.cell(row, 1, value=pred_date)
            ws.cell(row, 2, value="推荐")
            ws.cell(row, 30, value=p.get("pick",""))
            ws.cell(row, 33, value=p.get("confidence",""))
            ws.cell(row, 32, value=p.get("odds",""))
            ws.cell(row, 43, value=p.get("reason","")[:50])
            ws.cell(row, 44, value="自抓取")
            for c in range(1, 45):
                cell = ws.cell(row, c)
                try: cell.font = F; cell.fill = R1 if row%2==0 else R2; cell.alignment = C; cell.border = B
                except: pass
            added += 1
            changed = True
        print(f"  📊 推荐数据 +{added} 行")

    if changed:
        wb.save(MASTER_DB)
        print(f"  ✅ 总表已更新 (~/.hermes/data/)")
        return True
    else:
        print(f"  📊 无新增数据")
        return True


def write_to_obsidian(pred_date, matches_data, analysis_data):
    """写入Obsidian竞彩分析库"""
    # 确保目录结构
    for d in ["每日数据", "每日预测", "每日复盘"]:
        os.makedirs(f"{OBSIDIAN_VAULT}/{d}", exist_ok=True)

    # 写入每日数据
    data_path = f"{OBSIDIAN_VAULT}/每日数据/{pred_date}.md"
    with open(data_path, "w", encoding="utf-8") as f:
        f.write(f"# {pred_date} 原始比赛数据\n\n")
        f.write(f"> 自动抓取自 500.com | {now()}\n\n")
        if matches_data:
            f.write("| 场次 | 时间 | 联赛 | 主队 | 客队 | 主胜 | 平局 | 客胜 |\n")
            f.write("|:----:|:----:|:----:|:-----|:-----|:----:|:----:|:----:|\n")
            for m in matches_data:
                f.write(f"| {m.get('match_no','')} | {m.get('time','')} | {m.get('league','')} |")
                f.write(f" {m.get('home','')} | {m.get('away','')} |")
                f.write(f" {m.get('win_odds','')} | {m.get('draw_odds','')} | {m.get('lose_odds','')} |\n")
        f.write("\n---\n")
        f.write(f"\n*自动生成于 {now()}*\n")

    # 写入每日预测
    pred_path = f"{OBSIDIAN_VAULT}/每日预测/{pred_date}.md"
    with open(pred_path, "w", encoding="utf-8") as f:
        f.write(f"# {pred_date} 每日预测\n\n")
        f.write(f"> 自动抓取自 500.com | {now()}\n\n")
        f.write("\n---\n")
        f.write(f"\n*自动生成于 {now()}*\n")

    print(f"  ✅ 预测笔记已写入: {OBSIDIAN_VAULT}/每日预测/{pred_date}.md")
    return True


def write_review_to_obsidian(pred_date, review_data):
    """写入复盘结果到Obsidian"""
    os.makedirs(f"{OBSIDIAN_VAULT}/每日复盘", exist_ok=True)
    path = f"{OBSIDIAN_VAULT}/每日复盘/{pred_date}.md"
    with open(path, "w", encoding="utf-8") as f:
        f.write(f"# {pred_date} 每日复盘\n\n")
        f.write(f"> 自动更新于 {now()}\n\n")
        if review_data and isinstance(review_data, list):
            f.write("| 场次 | 对阵 | 推荐 | 比分 | 判断 |\n")
            f.write("|:----:|:-----|:----:|:----:|:----:|\n")
            for rv in review_data:
                f.write(f"| {rv.get('match_no','')} | {rv.get('home','')} vs {rv.get('away','')} |")
                f.write(f" {rv.get('pick','')} | {rv.get('score','')} | {rv.get('result','')} |\n")
    print(f"  ✅ 复盘笔记已写入: {path}")
    return True


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python3 daily_sync.py <all|matches|analysis|review> <日期> [数据JSON] [分析JSON]")
        sys.exit(1)

    action = sys.argv[1]
    pred_date = sys.argv[2]
    matches_data = json.loads(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[3] else []
    analysis_data = json.loads(sys.argv[4]) if len(sys.argv) > 4 and sys.argv[4] else {}

    print(f"🔄 每日同步 [{today()}] {pred_date}")
    print(f"   📋 比赛数据: {len(matches_data)} 场")

    if action in ("all", "matches"):
        print(f"  📊 → 写入总表Excel...")
        write_to_master_excel(pred_date, matches_data, analysis_data)
    elif action in ("all", "analysis"):
        print(f"  📝 → 写入竞彩分析库...")
        write_to_obsidian(pred_date, matches_data, analysis_data)

    # 同步到桌面
    if action in ("all", "matches"):
        try:
            subprocess.run(["cp", MASTER_DB, DESKTOP_DB], capture_output=True)
            print(f"  💻 桌面同步完成: {DESKTOP_DB}")
        except:
            print(f"  ⚠️ 桌面同步失败（可能被Windows锁文件）")

    print(f"✅ 每日同步完成 [{pred_date}]")
